#!/usr/bin/env python3
"""
Script to execute a batch of banking queries and save results to an Excel file.
Directly uses sql_generator functions instead of API calls.
"""
import pandas as pd
import time
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import logging
import sys
from pathlib import Path

# Add parent directory to path to import sql_generator
sys.path.append(str(Path(__file__).parent.parent))

# Import from sql_generator
from sql_generator import generate_sql, execute_query, get_db_engine

# Create output directory if it doesn't exist
output_dir = Path(__file__).parent / 'output'
output_dir.mkdir(exist_ok=True, parents=True)

# Configure logging
log_file=output_dir / 'query_execution.log'
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Initialize database engine
db_engine = None

# Style definitions for Excel
HEADER_FILL = PatternFill(
    start_color="4F81BD",  # Blue
    end_color="4F81BD",
    fill_type="solid"
)
HEADER_FONT = Font(color="FFFFFF", bold=True)
ERROR_FILL = PatternFill(
    start_color="FF0000",  # Red
    fill_type="solid"
)
WARNING_FILL = PatternFill(
    start_color="FFC000",  # Orange
    fill_type="solid"
)
SUCCESS_FILL = PatternFill(
    start_color="C6EFCE",  # Light green
    fill_type="solid"
)

def execute_single_query(query: str) -> Dict[str, Any]:
    """Execute a single query using sql_generator functions and return the results."""
    global db_engine
    
    if db_engine is None:
        try:
            db_engine = get_db_engine()
        except Exception as e:
            return {
                "query": query,
                "status": "error",
                "error": f"Failed to initialize database engine: {str(e)}",
                "sql": "",
                "results": []
            }
    
    try:
        # Generate SQL
        logger.info(f"Generating SQL for query: {query[:100]}...")
        sql = generate_sql(query)
        
        if not sql:
            return {
                "query": query,
                "status": "error",
                "error": "Failed to generate SQL",
                "sql": "",
                "results": []
            }
        
        logger.debug(f"Generated SQL: {sql}")
        
        # Execute the query
        logger.info("Executing SQL...")
        results = execute_query(sql, engine=db_engine, original_query=query)
        
        return {
            "query": query,
            "status": "success",
            "sql": sql,
            "results": results,
            "error": ""
        }
        
    except Exception as e:
        error_msg = str(e)
        logger.error(f"Error executing query: {error_msg}")
        return {
            "query": query,
            "status": "error",
            "error": error_msg,
            "sql": sql if 'sql' in locals() else "",
            "results": []
        }

def format_results_for_excel(results: List[Dict[str, Any]]) -> Dict:
    """Format query results for Excel output."""
    formatted = {
        "Query": [],
        "Status": [],
        "SQL Query": [],
        "Result Count": [],
        "Error": []
    }
    
    for result in results:
        formatted["Query"].append(result["query"])
        formatted["Status"].append(result["status"])
        formatted["SQL Query"].append(result.get("sql", ""))
        
        if result["status"] == "success":
            formatted["Result Count"].append(len(result["results"]))
            formatted["Error"].append("")
        else:
            formatted["Result Count"].append(0)
            formatted["Error"].append(result.get("error", "Unknown error"))
    
    return formatted

def save_to_excel(results: List[Dict[str, Any]], filename: str) -> None:
    """Save query results to an Excel file with formatting."""
    output_path = output_dir / filename
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Format summary sheet
    formatted = format_results_for_excel(results)
    df_summary = pd.DataFrame(formatted)
    
    # Write summary data
    for col_num, column_title in enumerate(df_summary.columns, 1):
        cell = ws_summary.cell(row=1, column=col_num, value=column_title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    
    for row_num, row_data in enumerate(df_summary.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws_summary.cell(row=row_num, column=col_num, value=cell_value)
    
    # Add conditional formatting for status
    for row in range(2, len(results) + 2):
        status_cell = ws_summary.cell(row=row, column=2)  # Status column
        if status_cell.value == "success":
            status_cell.fill = SUCCESS_FILL
        elif status_cell.value == "error":
            status_cell.fill = ERROR_FILL
        else:
            status_cell.fill = WARNING_FILL
    
    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.1
        ws_summary.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    # Create detailed results sheets
    for i, result in enumerate(results, 1):
        if result["status"] == "success" and result["results"]:
            # Create a new sheet for each successful query
            sheet_name = f"Query_{i}"
            if len(sheet_name) > 31:  # Excel sheet name limit
                sheet_name = sheet_name[:28] + "..."
            
            ws_detail = wb.create_sheet(title=sheet_name)
            
            # Write query and SQL
            ws_detail.append(["Query:", result["query"]])
            ws_detail.append(["Generated SQL:", result["sql"]])
            ws_detail.append([])  # Empty row
            
            # Write results as a table
            df_detail = pd.DataFrame(result["results"])
            
            # Write headers
            for col_num, column_title in enumerate(df_detail.columns, 1):
                cell = ws_detail.cell(row=4, column=col_num, value=column_title)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            
            # Write data
            for row_num, row_data in enumerate(df_detail.values, 5):
                for col_num, cell_value in enumerate(row_data, 1):
                    ws_detail.cell(row=row_num, column=col_num, value=cell_value)
            
            # Auto-adjust column widths
            for column in ws_detail.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.1
                ws_detail.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    # Save the workbook
    wb.save(output_path)
    logger.info(f"Results saved to {output_path}")

def main():
    # List of queries to execute

    queries = [
        "Give me female customers total deposit",
        "What is the net income for last quarter",
        "Give me current quarter interest income",
        "Give me total loans balance split by LCY and FCY"
        "Show me the total revenue across Kenya last year.",
        "Compare the split between LCY and FCY of Deposit for Kenya MTD.",
        "what is the last 6 month total income",
        "Give me depsoit and interest income for current month and last month",
        "Give me net income for last 12 months"
    ]
    
    logger.info(f"Starting execution of {len(queries)} queries...")
    
    # Execute all queries
    results = []
    for i, query in enumerate(queries, 1):
        logger.info(f"Processing query {i}/{len(queries)}: {query[:100]}...")
        try:
            result = execute_single_query(query)
            results.append(result)
            
            if result["status"] == "success":
                logger.info(f"  ✓ Success - {len(result['results'])} results")
            else:
                logger.error(f"  ✗ Error: {result.get('error', 'Unknown error')}")
            
            # Add a small delay between queries to avoid overwhelming the API
            time.sleep(1)
            
        except Exception as e:
            logger.error(f"Error processing query: {str(e)}")
            results.append({
                "query": query,
                "status": "error",
                "error": str(e),
                "sql": "",
                "results": []
            })
    
    # Save results to Excel
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"query_results_{timestamp}.xlsx"
    save_to_excel(results, output_file)
    
    # Print summary
    success_count = sum(1 for r in results if r["status"] == "success")
    error_count = len(results) - success_count
    
    logger.info("\n" + "="*50)
    logger.info(f"Execution completed. Success: {success_count}, Errors: {error_count}")
    logger.info(f"Results saved to: {output_file}")
    logger.info("="*50)

if __name__ == "__main__":
    main()
